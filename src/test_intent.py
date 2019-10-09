import json, urllib3, collections, copy
from openpyxl import load_workbook
import pandas as pd
import numpy as np


def post_etri(text):
    requestJson = {
        "access_key" : "4ee51c5e-7d13-4f91-9516-5f68c4fe26f3",
        'argument' : {
            "text" : text,
            "analysis_code" : "dparse"
            }
        }

    http = urllib3.PoolManager()
    response = http.request(
        "POST",
        "http://aiopen.etri.re.kr:8000/WiseNLU",
        headers={
            "Content-Type":"application/json; charset=UTF-8"}, 
        body = json.dumps(requestJson)
        )
    data = json.loads(response.data.decode("utf-8"))
    dep = data["return_object"]["sentence"][0]["dependency"]
    ner = data["return_object"]["sentence"][0]["NE"]
    return dep, ner


def post_mecab(text):
    http = urllib3.PoolManager()
    response = http.request(
        "POST",
        "https://www.officelog.net/mecab",
        headers={
            "Content-Type":"application/json; charset=UTF-8"}, 
        body = json.dumps({"input": text})
        )
    mmm = json.loads(response.data.decode("utf-8"))["output"]
    return mmm


def load_tag():
    """
    tag를 로딩 함
    """
    wb = load_workbook("../doc/config.xlsx")
    ws = wb["tag"]

    tag_idx = dict()
    tag_dic = dict()
    for _, row in enumerate(ws.rows, start=1):
        for c_idx, cell in enumerate(row, start=1):
            if cell.value is not None:
                tag_idx[c_idx] = cell.value.strip()
                tag_dic[cell.value.strip()] = ""
        break

    tag_list = []
    for r_idx, row in enumerate(ws.rows, start=1):
        if r_idx < 2: continue
        tag_ok = False
        tag_row = copy.deepcopy(tag_dic)
        for c_idx, cell in enumerate(row, start=1):
            if cell.value is not None and c_idx in tag_idx:
                tag = tag_idx[c_idx]
                tag_row[tag] = cell.value.strip()
                tag_ok = True
        if tag_ok:
            tag_list.append(tag_row)
    
    return tag_list


def get_intents(tag_list, dep, ner, morp):
    """
    tag를 match 하여 intent를 파악 함
    """
    intents = collections.OrderedDict()
    for tag_row in tag_list:
        match = dict()
        for tag_0, val_0 in tag_row.items():
            if tag_0 == "INTENT" or tag_0 == "VALUE" or val_0 == "": continue
            if tag_0 in match and match[tag_0] is not None: continue
            match[tag_0] = None
            for mmm in morp:
                for val_1, tag_1 in mmm.items():
                    if not match[tag_0] is not None and tag_0 == tag_1:
                        if val_0 == val_1: match[tag_0] = val_1
                        else: match[tag_0] = None
        result = True
        value = None
        for tag_0, mat_0 in match.items():
            if not mat_0:
                result = False
                break
            else:
                value = mat_0
        if result and value:
            if tag_row["VALUE"] == "O":
                intents[tag_row["INTENT"]] = value
            else:
                intents["INTENT"] = tag_row["INTENT"]
    
    if "INTENT" in intents:
        for ne in ner:
            intents[ne["type"]] = ne["text"]
    return intents


def match_excel_tag():
    """
    excel tag를 match 하여 저장 함
    """
    tag_list = load_tag()

    wb = load_workbook("../doc/config2.xlsx")
    ws = wb["example"]
    for index, row in enumerate(ws.rows, start=1):
        if 1 < index:
            text = row[1].value.strip()
            dep, ner = post_etri(text)
            morp = post_mecab(text)
            intents = get_intents(tag_list, dep, ner, morp)
            intents = list(intents.items())
            print(text, intents)
            ws.cell(row=index, column=3).value = str(intents)

    wb.save("../doc/config2.xlsx")


def match_tag(text):
    """
    한개의 tag를 match 해 봄
    """
    tag_list = load_tag()
    dep, ner = post_etri(text)
    morp = post_mecab(text)
    # print(dep)
    # print(ner)
    # print(morp)
    print(list(get_intents(tag_list, dep, ner, morp).items()))


if __name__ == "__main__":
    # match_tag("10월 12일에 갈만한 행사 찾아줘")
    match_excel_tag()
    

    